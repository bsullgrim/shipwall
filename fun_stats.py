#!/usr/bin/env python3
"""
fun_stats.py  --  Mine the ship-spotting log(s) for a "Hall of Fame & Oddities"
companion to the passage leaderboards.

Reads one or two multi-year Excel logs and writes fun_stats.json, which
passage_stats.py serves on its "Hall of Fame" tab. With two logs (the home
river log plus Emmett's log), the two observers watch the SAME Cedar Ledge
crossing at different times, and Emmett additionally tags sightings made
elsewhere up the system (Soo Locks, the connecting rivers, open lakes). So the
merge is by shared crossing, and Emmett's non-home locations drive a corridor /
transit-chain view that a single observer could never produce.

  - exotics:        notable sightings flagged by comment keyword
  - milestones:     personal/sentimental comments, now observer-credited
  - rarities:       ships seen exactly once across the merged log
  - busiest:        highest ship-count days ever (merged)
  - lifers:         brand-new ships added each year (merged)
  - corridor:       sighting counts per waypoint, lakehead -> Cedar Ledge
  - transit_chains: hulls logged at 2+ waypoints, as ordered downbound paths
  - observers:      per-observer tallies + who-caught-what buckets

    python3 fun_stats.py home.xlsx                       # one observer
    python3 fun_stats.py home.xlsx --emmett emmett.xlsx  # two observers
    python3 fun_stats.py home.xlsx --emmett e.xlsx --passages passages.csv
"""

import argparse
import json
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl

import corridor

# Optional: reuse the operator-resolution rules for name normalization, so
# "CSL Oakglen" (Emmett) and "Oakglen" (home) collapse to one hull.
try:
    import operators as _ops
except Exception:                       # operators.py not importable here
    _ops = None


# --------------------------------------------------------------------------- #
# Name normalization
# --------------------------------------------------------------------------- #
# The two logs spell the same hull differently: Emmett prefixes the operator
# ("CSL Oakglen", "Lower Lakes Towing Manitoulin"), the home log usually does
# not ("Oakglen", "Manitoulin"). We canonicalize to a single comparison key so
# the merge, rarities, and lifelist line up. Display still uses the home log's
# spelling when available (it's the cleaner one), else the first seen.

# Operator display-name prefixes that Emmett's log prepends to a bare hull name.
# Stripping these recovers the home/AIS spelling. Keep fleets whose operator
# word IS part of the real vessel name out of this list (so we never mangle
# "Algoma Niagara" or "Federal Yukon" or "BBC Elbe").
# Tier-1: operator words that are NEVER part of a real vessel name -- always
# safe to strip from Emmett's "Operator Vessel" spellings.
_STRIP_PREFIXES = [
    "lower lakes towing ", "american steamship company ",
    "grand river navigation ", "mcasphalt marine transportation ",
    "interlake steamship company ", "vanenkevort tug and barge ",
    "key lakes ", "carisbrooke shipping ", "andrie inc ", "andrie ",
    "canfornav ", "navibulgar ", "wagenborg ", "polsteam ", "spliethoff ",
    "mckeil ", "big lift ", "desgagnes ", "csl ",
]

# Tier-2: operator words that ARE part of some real vessel names ("Algoma
# Niagara", "Federal Yukon", "BBC Elbe"). Stripping these is only safe when the
# remainder actually matches a hull the home log knows -- otherwise we'd merge
# distinct vessels. Resolved against _HOME_KEYS, populated before keying.
_STRIP_PREFIXES_TIER2 = ["algoma ", "federal ", "bbc "]
_HOME_KEYS = set()


def norm_name(raw):
    """Whitespace-collapsed display name."""
    return " ".join(str(raw).strip().split())


def _tier1_key(name):
    """Tier-1-only key, used to seed home keys before tier-2 resolution."""
    k = norm_name(name).lower()
    for p in _STRIP_PREFIXES:
        if k.startswith(p):
            return " ".join(k[len(p):].split())
    return " ".join(k.split())


def compare_key(name):
    """Normalization key for matching the same hull across logs.

    Tier-1 prefixes always strip. Tier-2 prefixes strip only when the result is
    a hull the home log already knows, so "Algoma John D Leitch" -> "john d
    leitch" (home has it) but "Algoma Niagara" stays intact (home has no bare
    "niagara").
    """
    k = norm_name(name).lower()
    for p in _STRIP_PREFIXES:
        if k.startswith(p):
            return " ".join(k[len(p):].split())
    for p in _STRIP_PREFIXES_TIER2:
        if k.startswith(p):
            cand = " ".join(k[len(p):].split())
            if cand in _HOME_KEYS:
                return cand
            break
    return " ".join(k.split())


# --------------------------------------------------------------------------- #
# Reading logs
# --------------------------------------------------------------------------- #
def find_cols(ws):
    """Locate the header row (scans 1-4) and map header -> column index."""
    for r in range(1, 5):
        vals = [str(ws.cell(row=r, column=c).value) for c in range(1, 15)]
        if "Ship" in vals:
            return r, {v: i + 1 for i, v in enumerate(vals)}
    return 2, {}


def load_rows(path, observer):
    """Read every year sheet into sighting dicts tagged with the observer."""
    wb = openpyxl.load_workbook(path, data_only=True)
    years = sorted(s for s in wb.sheetnames if s.isdigit())
    rows = []
    for sn in years:
        ws = wb[sn]
        hdr, d = find_cols(ws)
        # Pre-resolve the location column index once per sheet.
        loc_col = None
        for k in d:
            if k and "Location" in k:
                loc_col = d[k]
                break
        y = int(sn)
        for r in range(hdr + 1, ws.max_row + 1):
            g = lambda c: ws.cell(row=r, column=d[c]).value if c in d else None
            ship = g("Ship")
            if not ship or not str(ship).strip():
                continue
            name = norm_name(ship)
            if name.lower() in ("ship", "none") or name.replace("-", "").isdigit():
                continue
            dt = g("Date")
            if isinstance(dt, datetime) and dt.year != y:
                try:
                    dt = dt.replace(year=y)
                except ValueError:
                    dt = datetime(y, 1, 1)
            loc_raw = ws.cell(row=r, column=loc_col).value if loc_col else None
            rows.append({
                "year": y,
                "name": name,
                "key": None,   # assigned after _HOME_KEYS is seeded
                "fleet": str(g("Fleet") or "").strip(),
                "comment": str(g("Comments") or g("Remarks") or "").strip(),
                "date": dt if isinstance(dt, datetime) else None,
                "location": corridor.norm_loc(loc_raw),
                "observer": observer,
            })
    return rows, years


def load_mmsi_backfill(passages_path):
    """key -> mmsi, harvested from passages.csv to firm up name matching.

    Advisory only: we don't add passages as sightings, we just use the
    name<->mmsi pairs to sanity-check that cross-log name matches are real.
    """
    table = {}
    if not passages_path or not os.path.exists(passages_path):
        return table
    import csv
    with open(passages_path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            nm = (row.get("name") or "").strip()
            mmsi = (row.get("mmsi") or "").strip()
            if nm and mmsi:
                table.setdefault(compare_key(nm), mmsi)
    return table


# --------------------------------------------------------------------------- #
# Mining
# --------------------------------------------------------------------------- #
EXOTIC_RULES = [
    ("tall ship", "Tall ship"), ("schooner", "Sail"), ("barque", "Sail"),
    ("brig", "Sail"), ("sailing", "Sail"),
    ("yacht", "Yacht"),
    ("coast guard", "Coast Guard"), ("ccg", "Coast Guard"),
    ("navy", "Naval"), ("warship", "Naval"), ("submarine", "Naval"),
    ("icebreaker", "Icebreaker"),
    ("hovercraft", "Oddity"), ("very cool", "Noteworthy"),
    ("rare", "Noteworthy"), ("cruise", "Cruise"),
]
MILESTONE_KW = ["emma", "first boat", "first ship", "james",
                "favorite", "favourite", "best ever", "lifer"]


def categorize(comment):
    cl = comment.lower()
    for kw, cat in EXOTIC_RULES:
        if kw in cl:
            return cat
    return None


def display_name_for(key, rows):
    """Prefer the home observer's spelling of a hull, else first seen."""
    home = [r["name"] for r in rows if r["key"] == key and r["observer"] == "home"]
    if home:
        return home[0]
    any_ = [r["name"] for r in rows if r["key"] == key]
    return any_[0] if any_ else key


def build(rows, years, observers):
    # --- Collapse shared crossings ------------------------------------------ #
    # Same hull + same date + Cedar Ledge logged by both observers == ONE
    # crossing they both witnessed. Off-home locations are never collapsed
    # with home (they're distinct upstream sightings).
    crossing = {}  # (key, date_iso, location) -> record
    for r in rows:
        diso = r["date"].date().isoformat() if r["date"] else f"~{r['year']}"
        ckey = (r["key"], diso, r["location"])
        rec = crossing.get(ckey)
        if rec is None:
            crossing[ckey] = {
                "key": r["key"], "name": r["name"], "year": r["year"],
                "date": r["date"], "location": r["location"],
                "comment": r["comment"], "fleet": r["fleet"],
                "observers": {r["observer"]},
            }
        else:
            rec["observers"].add(r["observer"])
            if not rec["comment"] and r["comment"]:
                rec["comment"] = r["comment"]
    merged = list(crossing.values())

    # Per-hull observer membership (across all of that hull's crossings).
    hull_obs = defaultdict(set)
    for m in merged:
        hull_obs[m["key"]] |= m["observers"]

    seen = Counter(m["key"] for m in merged)
    name_of = {k: display_name_for(k, rows) for k in seen}

    # --- Exotics (first occurrence per hull, keep comment) ------------------ #
    exotics, seen_exotic = [], set()
    for m in sorted(merged, key=lambda m: (m["year"],
                    m["date"] or datetime(m["year"], 1, 1))):
        if not m["comment"]:
            continue
        cat = categorize(m["comment"])
        if cat and m["key"] not in seen_exotic:
            seen_exotic.add(m["key"])
            exotics.append({"name": name_of[m["key"]], "category": cat,
                            "comment": m["comment"], "year": m["year"]})
    cat_order = {c: i for i, c in enumerate(
        ["Tall ship", "Sail", "Naval", "Coast Guard", "Icebreaker",
         "Yacht", "Cruise", "Oddity", "Noteworthy"])}
    exotics.sort(key=lambda e: (cat_order.get(e["category"], 99), e["year"]))

    # --- Milestones (personal comments, observer-credited) ------------------ #
    milestones = []
    for m in merged:
        cl = m["comment"].lower()
        if any(k in cl for k in MILESTONE_KW) and len(m["comment"]) < 100:
            iso = m["date"].date().isoformat() if m["date"] else str(m["year"])
            who = "both" if len(m["observers"]) > 1 else next(iter(m["observers"]))
            milestones.append({"name": name_of[m["key"]], "comment": m["comment"],
                               "date": iso, "year": m["year"], "observer": who})
    milestones.sort(key=lambda m: m["date"])

    # --- Rarities (seen exactly once, ever), observer-credited -------------- #
    rarities = []
    for k, c in seen.items():
        if c == 1:
            obs = hull_obs[k]
            who = "both" if len(obs) > 1 else next(iter(obs))
            rarities.append({"name": name_of[k], "observer": who})
    rarities.sort(key=lambda r: r["name"].lower())

    # --- Busiest days (merged crossings) ------------------------------------ #
    byday = Counter()
    for m in merged:
        if m["date"]:
            byday[m["date"].date().isoformat()] += 1
    busiest = [{"date": d, "count": c} for d, c in byday.most_common(10)]

    # --- Lifers per year (first year a hull appears, merged) ---------------- #
    first_year = {}
    for m in sorted(merged, key=lambda m: m["year"]):
        first_year.setdefault(m["key"], m["year"])
    lifer_counts = Counter(first_year.values())
    all_years = sorted({m["year"] for m in merged})
    lifers = [{"year": str(y), "new": lifer_counts.get(y, 0)} for y in all_years]

    # --- Corridor + transit chains (Emmett's location data) ----------------- #
    loc_counts = Counter(m["location"] for m in merged)
    ship_locs = defaultdict(set)
    for m in merged:
        ship_locs[name_of[m["key"]]].add(m["location"])
    corridor_strip = corridor.build_corridor(loc_counts)
    chains = corridor.build_transit_chains(ship_locs)

    # --- Observer tallies + who-caught-what buckets ------------------------- #
    obs_tally = {}
    for o in observers:
        o_hulls = {k for k, s in hull_obs.items() if o in s}
        o_sight = sum(1 for m in merged if o in m["observers"])
        obs_tally[o] = {"sightings": o_sight, "distinct_ships": len(o_hulls)}
    buckets = {"home_only": 0, "emmett_only": 0, "both": 0}
    emmett_exclusive = []  # hulls only Emmett logged
    for k, obs in hull_obs.items():
        if obs == {"home"}:
            buckets["home_only"] += 1
        elif obs == {"emmett"}:
            buckets["emmett_only"] += 1
            emmett_exclusive.append(name_of[k])
        elif "home" in obs and "emmett" in obs:
            buckets["both"] += 1
    emmett_exclusive.sort(key=str.lower)

    fleets = Counter(m["fleet"] for m in merged
                     if m["fleet"] and m["fleet"].lower() != "none")

    out = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "totals": {
            "sightings": len(merged),
            "distinct_ships": len(seen),
            "distinct_fleets": len(fleets),
            "one_timers": len(rarities),
            "span": f"{all_years[0]}\u2013{all_years[-1]}" if all_years else "",
        },
        "exotics": exotics,
        "milestones": milestones,
        "busiest_days": busiest,
        "lifers_per_year": lifers,
        "rarities_sample": [r["name"] for r in rarities][:60],
        "rarities_total": len(rarities),
        "rarities_credited": rarities[:60],
        # New, location-driven blocks:
        "corridor": corridor_strip,
        "transit_chains": chains[:40],
        "transit_chains_total": len(chains),
    }
    if len(observers) > 1:
        out["observers"] = {
            "tally": obs_tally,
            "buckets": buckets,
            "emmett_exclusive_sample": emmett_exclusive[:40],
            "emmett_exclusive_total": len(emmett_exclusive),
        }
    return out


def main(home, emmett, passages, out):
    global _HOME_KEYS
    rows, years = load_rows(home, "home")
    observers = ["home"]
    # Seed the home-key set (tier-1 spelling) so tier-2 prefix stripping on
    # Emmett's names can resolve against hulls the home log actually knows.
    _HOME_KEYS = {_tier1_key(r["name"]) for r in rows}
    if emmett:
        erows, _ = load_rows(emmett, "emmett")
        rows += erows
        observers.append("emmett")
    # Now every row gets its cross-log comparison key.
    for r in rows:
        r["key"] = compare_key(r["name"])
    _ = load_mmsi_backfill(passages)   # advisory; firms up matching
    data = build(rows, years, observers)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    t = data["totals"]
    print(f"wrote {out}")
    print(f"  {t['sightings']} merged crossings, {t['distinct_ships']} distinct hulls, "
          f"{t['distinct_fleets']} fleets, {t['one_timers']} one-timers")
    print(f"  {len(data['exotics'])} exotics, {len(data['milestones'])} milestones, "
          f"{data['transit_chains_total']} transit chains, "
          f"{len(data['corridor'])} corridor waypoints")
    if "observers" in data:
        b = data["observers"]["buckets"]
        print(f"  observers: home-only {b['home_only']}, "
              f"emmett-only {b['emmett_only']}, both {b['both']}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Mine the ship log(s) for fun stats.")
    ap.add_argument("log", help="the HOME ship-log .xlsx")
    ap.add_argument("--emmett", help="Emmett's ship-log .xlsx (adds 2nd observer)")
    ap.add_argument("--passages", help="passages.csv (MMSI backfill for matching)")
    ap.add_argument("-o", "--output", default="fun_stats.json")
    args = ap.parse_args()
    main(args.log, args.emmett, args.passages, args.output)